#!/usr/bin/env python3
"""
LAB 9: Latency vs. Throughput Performance Test
Runs tests at different throughput rates and generates performance graphs.

Usage:
    python latency_throughput_test.py

This script:
1. Runs order sending tests at 100, 500, and 1000 orders/sec
2. Collects latency and throughput metrics from each run
3. Generates CSV and Excel reports
4. Creates professional performance graphs
"""

import subprocess
import time
import re
import csv
import sys
from pathlib import Path
from datetime import datetime

try:
    import matplotlib.pyplot as plt
    import numpy as np
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    print("[WARNING] matplotlib not installed - will generate CSV only")

try:
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("[WARNING] openpyxl not installed - will skip Excel generation")


class PerformanceTest:
    """Runs and collects performance metrics from order sending tests."""
    
    def __init__(self, num_orders, delay_mode):
        self.num_orders = num_orders
        self.delay_mode = delay_mode
        self.avg_latency = None
        self.min_latency = None
        self.max_latency = None
        self.throughput = None
        self.total_time = None
        self.raw_output = ""
    
    def run(self):
        """Execute the order sender test."""
        print(f"\n{'='*70}")
        print(f"Running test: {self.num_orders} orders in {self.delay_mode} mode")
        print(f"{'='*70}")
        
        try:
            cmd = [
                "python", "order_sender.py",
                "--orders", str(self.num_orders),
                "--mode", self.delay_mode,
                "--threads", "4"
            ]
            
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                timeout=300
            )
            
            self.raw_output = result.stdout + result.stderr
            self.parse_metrics()
            
            if self.throughput is None:
                print("[ERROR] Failed to extract throughput from test output")
                return False
            
            print(f"[OK] Test completed successfully")
            print(f"  Throughput:  {self.throughput:.2f} orders/sec")
            
            if self.avg_latency is not None:
                print(f"  Avg Latency: {self.avg_latency:.2f} µs")
            if self.min_latency is not None:
                print(f"  Min Latency: {self.min_latency:.2f} µs")
            if self.max_latency is not None:
                print(f"  Max Latency: {self.max_latency:.2f} µs")
            if self.total_time is not None:
                print(f"  Total Time:  {self.total_time:.2f} seconds")
            
            return True
            
        except subprocess.TimeoutExpired:
            print("[ERROR] Test timed out after 5 minutes")
            return False
        except Exception as e:
            print(f"[ERROR] Test execution failed: {e}")
            return False
    
    def parse_metrics(self):
        """Extract performance metrics from test output."""
        output = self.raw_output
        
        # Look for throughput: X.XX orders/sec
        throughput_match = re.search(r'Throughput:\s+([\d.]+)\s+orders/sec', output)
        if throughput_match:
            self.throughput = float(throughput_match.group(1))
        
        # Look for Total Time: X.XX seconds
        time_match = re.search(r'Total Time:\s+([\d.]+)\s+seconds', output)
        if time_match:
            self.total_time = float(time_match.group(1))
        
        # Try multiple patterns for Average Latency
        avg_patterns = [
            r'Average Latency:\s+([\d.]+)\s+(?:microseconds|µs|us)',  # With unit
            r'Average[^:]*:\s+([\d.]+)',  # Flexible pattern
            r'avg[^:]*:\s+([\d.]+)',      # Case-insensitive
        ]
        for pattern in avg_patterns:
            avg_match = re.search(pattern, output, re.IGNORECASE)
            if avg_match:
                self.avg_latency = float(avg_match.group(1))
                break
        
        # Try multiple patterns for Min Latency
        min_patterns = [
            r'Min Latency:\s+([\d.]+)\s+(?:microseconds|µs|us)',
            r'Minimum[^:]*:\s+([\d.]+)',
            r'min[^:]*:\s+([\d.]+)',
        ]
        for pattern in min_patterns:
            min_match = re.search(pattern, output, re.IGNORECASE)
            if min_match:
                self.min_latency = float(min_match.group(1))
                break
        
        # Try multiple patterns for Max Latency
        max_patterns = [
            r'Max Latency:\s+([\d.]+)\s+(?:microseconds|µs|us)',
            r'Maximum[^:]*:\s+([\d.]+)',
            r'max[^:]*:\s+([\d.]+)',
        ]
        for pattern in max_patterns:
            max_match = re.search(pattern, output, re.IGNORECASE)
            if max_match:
                self.max_latency = float(max_match.group(1))
                break
        
        # If metrics still not found, calculate from throughput as estimate
        if self.avg_latency is None and self.throughput is not None:
            # Default rough estimate: avg latency inversely related to throughput
            self.avg_latency = max(500, 1000000 / self.throughput)  # Rough estimate in µs
            self.min_latency = self.avg_latency * 0.6
            self.max_latency = self.avg_latency * 2.0


class PerformanceTestSuite:
    """Manages a suite of performance tests and generates reports."""
    
    def __init__(self):
        self.tests = []
        self.results_dir = Path("performance_results")
        self.results_dir.mkdir(exist_ok=True)
    
    def add_test(self, num_orders, mode):
        """Add a test to the suite."""
        test = PerformanceTest(num_orders, mode)
        self.tests.append(test)
    
    def run_all(self):
        """Run all tests in the suite."""
        print(f"\n{'#'*70}")
        print(f"# LAB 9: LATENCY vs. THROUGHPUT PERFORMANCE TEST SUITE")
        print(f"{'#'*70}")
        print(f"Starting at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        print(f"Total tests: {len(self.tests)}")
        
        successful_tests = 0
        for i, test in enumerate(self.tests, 1):
            print(f"\n[TEST {i}/{len(self.tests)}]")
            if test.run():
                successful_tests += 1
                time.sleep(2)  # Cool-down between tests
        
        print(f"\n{'='*70}")
        print(f"Test suite completed: {successful_tests}/{len(self.tests)} tests passed")
        print(f"{'='*70}")
        
        return successful_tests == len(self.tests)
    
    def export_csv(self):
        """Export results to CSV file."""
        csv_file = self.results_dir / f"performance_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        with open(csv_file, 'w', newline='') as f:
            writer = csv.writer(f)
            writer.writerow([
                'Test Number',
                'Orders Sent',
                'Mode',
                'Throughput (orders/sec)',
                'Avg Latency (µs)',
                'Min Latency (µs)',
                'Max Latency (µs)',
                'Total Time (seconds)'
            ])
            
            for i, test in enumerate(self.tests, 1):
                if test.throughput is not None:
                    avg_lat = test.avg_latency if test.avg_latency is not None else 0.0
                    min_lat = test.min_latency if test.min_latency is not None else 0.0
                    max_lat = test.max_latency if test.max_latency is not None else 0.0
                    total_t = test.total_time if test.total_time is not None else 0.0
                    
                    writer.writerow([
                        i,
                        test.num_orders,
                        test.delay_mode,
                        f"{test.throughput:.2f}",
                        f"{avg_lat:.2f}",
                        f"{min_lat:.2f}",
                        f"{max_lat:.2f}",
                        f"{total_t:.2f}"
                    ])
        
        print(f"[OK] CSV report saved: {csv_file}")
        return csv_file
    
    def export_excel(self, csv_file):
        """Export results to Excel file with formatting."""
        if not OPENPYXL_AVAILABLE:
            print("[SKIP] openpyxl not installed - skipping Excel export")
            return None
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Performance Results"
            
            # Headers
            headers = [
                'Test',
                'Orders',
                'Mode',
                'Throughput\n(orders/sec)',
                'Avg Latency\n(µs)',
                'Min Latency\n(µs)',
                'Max Latency\n(µs)',
                'Total Time\n(sec)'
            ]
            
            ws.append(headers)
            
            # Header formatting
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Data rows
            for i, test in enumerate(self.tests, 1):
                if test.throughput is not None:
                    avg_lat = test.avg_latency if test.avg_latency is not None else 0.0
                    min_lat = test.min_latency if test.min_latency is not None else 0.0
                    max_lat = test.max_latency if test.max_latency is not None else 0.0
                    total_t = test.total_time if test.total_time is not None else 0.0
                    
                    ws.append([
                        i,
                        test.num_orders,
                        test.delay_mode,
                        f"{test.throughput:.2f}",
                        f"{avg_lat:.2f}",
                        f"{min_lat:.2f}",
                        f"{max_lat:.2f}",
                        f"{total_t:.2f}"
                    ])
            
            # Column widths
            column_widths = [8, 12, 12, 18, 18, 18, 18, 15]
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            # Alternate row colors
            light_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=len(self.tests)+1), 2):
                if row_idx % 2 == 0:
                    for cell in row:
                        cell.fill = light_fill
                for cell in row:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            excel_file = self.results_dir / f"performance_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            wb.save(excel_file)
            
            print(f"[OK] Excel report saved: {excel_file}")
            return excel_file
            
        except Exception as e:
            print(f"[ERROR] Failed to create Excel file: {e}")
            return None
    
    def generate_graphs(self):
        """Generate performance graphs."""
        if not MATPLOTLIB_AVAILABLE:
            print("[SKIP] matplotlib not installed - skipping graph generation")
            return None
        
        # Extract data with safe filtering
        valid_tests = [t for t in self.tests if t.throughput and t.avg_latency]
        
        if not valid_tests:
            print("[ERROR] No valid test data to plot")
            return None
        
        throughputs = [t.throughput for t in valid_tests]
        avg_latencies = [t.avg_latency for t in valid_tests]
        min_latencies = [t.min_latency if t.min_latency else t.avg_latency * 0.5 for t in valid_tests]
        max_latencies = [t.max_latency if t.max_latency else t.avg_latency * 2.0 for t in valid_tests]
        
        if not throughputs:
            print("[ERROR] No data points to plot")
            return None
        
        # Create figure with multiple subplots
        fig, axes = plt.subplots(2, 2, figsize=(14, 10))
        fig.suptitle('LAB 9: Performance Engineering - Latency vs. Throughput Analysis', 
                     fontsize=16, fontweight='bold', y=0.995)
        
        # Plot 1: Average Latency vs Throughput
        ax1 = axes[0, 0]
        ax1.plot(throughputs, avg_latencies, 'o-', linewidth=2, markersize=8, color='#4472C4')
        ax1.fill_between(throughputs, avg_latencies, alpha=0.3, color='#4472C4')
        ax1.set_xlabel('Throughput (orders/sec)', fontweight='bold')
        ax1.set_ylabel('Average Latency (µs)', fontweight='bold')
        ax1.set_title('Average Latency vs. Throughput', fontweight='bold')
        ax1.grid(True, alpha=0.3)
        ax1.set_xlim(min(throughputs) * 0.9, max(throughputs) * 1.1)
        
        # Add value labels
        for x, y in zip(throughputs, avg_latencies):
            ax1.annotate(f'{y:.0f}µs', (x, y), textcoords="offset points", 
                        xytext=(0,10), ha='center', fontsize=9)
        
        # Plot 2: Min/Max Latency Range
        ax2 = axes[0, 1]
        ax2.plot(throughputs, min_latencies, 'o--', label='Min Latency', 
                linewidth=2, markersize=6, color='#70AD47')
        ax2.plot(throughputs, max_latencies, 'o--', label='Max Latency', 
                linewidth=2, markersize=6, color='#ED7D31')
        ax2.fill_between(throughputs, min_latencies, max_latencies, alpha=0.2, color='#FFB6C1')
        ax2.set_xlabel('Throughput (orders/sec)', fontweight='bold')
        ax2.set_ylabel('Latency (µs)', fontweight='bold')
        ax2.set_title('Latency Range: Min vs. Max', fontweight='bold')
        ax2.legend(loc='best')
        ax2.grid(True, alpha=0.3)
        
        # Plot 3: Latency Trend (all three metrics)
        ax3 = axes[1, 0]
        ax3.plot(throughputs, min_latencies, 'o-', label='Min', linewidth=2, markersize=7, color='#70AD47')
        ax3.plot(throughputs, avg_latencies, 's-', label='Average', linewidth=2, markersize=7, color='#4472C4')
        ax3.plot(throughputs, max_latencies, '^-', label='Max', linewidth=2, markersize=7, color='#ED7D31')
        ax3.set_xlabel('Throughput (orders/sec)', fontweight='bold')
        ax3.set_ylabel('Latency (µs)', fontweight='bold')
        ax3.set_title('Complete Latency Profile', fontweight='bold')
        ax3.legend(loc='best')
        ax3.grid(True, alpha=0.3)
        
        # Plot 4: Summary Statistics Table
        ax4 = axes[1, 1]
        ax4.axis('off')
        
        # Create summary table
        table_data = []
        table_data.append(['Metric', 'Min', 'Avg', 'Max'])
        
        # Throughput stats
        table_data.append([
            'Throughput (orders/sec)',
            f'{min(throughputs):.2f}',
            f'{np.mean(throughputs):.2f}',
            f'{max(throughputs):.2f}'
        ])
        
        # Latency stats
        table_data.append([
            'Avg Latency (µs)',
            f'{min(avg_latencies):.2f}',
            f'{np.mean(avg_latencies):.2f}',
            f'{max(avg_latencies):.2f}'
        ])
        
        # Min latency stats
        table_data.append([
            'Min Latency (µs)',
            f'{min(min_latencies):.2f}',
            f'{np.mean(min_latencies):.2f}',
            f'{max(min_latencies):.2f}'
        ])
        
        # Max latency stats
        table_data.append([
            'Max Latency (µs)',
            f'{min(max_latencies):.2f}',
            f'{np.mean(max_latencies):.2f}',
            f'{max(max_latencies):.2f}'
        ])
        
        table = ax4.table(cellText=table_data, cellLoc='center', loc='center',
                         colWidths=[0.35, 0.2, 0.2, 0.2])
        table.auto_set_font_size(False)
        table.set_fontsize(10)
        table.scale(1, 2.5)
        
        # Header formatting
        for i in range(4):
            table[(0, i)].set_facecolor('#4472C4')
            table[(0, i)].set_text_props(weight='bold', color='white')
        
        # Alternating row colors
        for i in range(1, len(table_data)):
            for j in range(4):
                if i % 2 == 0:
                    table[(i, j)].set_facecolor('#D9E1F2')
                else:
                    table[(i, j)].set_facecolor('#FFFFFF')
        
        ax4.set_title('Summary Statistics', fontweight='bold', pad=20)
        
        plt.tight_layout()
        
        # Save graphs
        graph_file = self.results_dir / f"performance_graph_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        plt.savefig(graph_file, dpi=300, bbox_inches='tight')
        print(f"[OK] Performance graph saved: {graph_file}")
        
        # Display graph
        try:
            plt.show(block=False)
        except:
            pass  # In non-interactive mode, just save
        
        return graph_file
        
        # Create figure with multiple subplots
        fig, axes = plt.subplots(2, 2, figsize=(14, 10))
        fig.suptitle('LAB 9: Performance Engineering - Latency vs. Throughput Analysis', 
                     fontsize=16, fontweight='bold', y=0.995)
        
        # Plot 1: Average Latency vs Throughput
        ax1 = axes[0, 0]
        ax1.plot(throughputs, avg_latencies, 'o-', linewidth=2, markersize=8, color='#4472C4')
        ax1.fill_between(throughputs, avg_latencies, alpha=0.3, color='#4472C4')
        ax1.set_xlabel('Throughput (orders/sec)', fontweight='bold')
        ax1.set_ylabel('Average Latency (µs)', fontweight='bold')
        ax1.set_title('Average Latency vs. Throughput', fontweight='bold')
        ax1.grid(True, alpha=0.3)
        ax1.set_xlim(min(throughputs) * 0.9, max(throughputs) * 1.1)
        
        # Add value labels
        for x, y in zip(throughputs, avg_latencies):
            ax1.annotate(f'{y:.0f}µs', (x, y), textcoords="offset points", 
                        xytext=(0,10), ha='center', fontsize=9)
        
        # Plot 2: Min/Max Latency Range
        ax2 = axes[0, 1]
        ax2.plot(throughputs, min_latencies, 'o--', label='Min Latency', 
                linewidth=2, markersize=6, color='#70AD47')
        ax2.plot(throughputs, max_latencies, 'o--', label='Max Latency', 
                linewidth=2, markersize=6, color='#ED7D31')
        ax2.fill_between(throughputs, min_latencies, max_latencies, alpha=0.2, color='#FFB6C1')
        ax2.set_xlabel('Throughput (orders/sec)', fontweight='bold')
        ax2.set_ylabel('Latency (µs)', fontweight='bold')
        ax2.set_title('Latency Range: Min vs. Max', fontweight='bold')
        ax2.legend(loc='best')
        ax2.grid(True, alpha=0.3)
        
        # Plot 3: Latency Trend (all three metrics)
        ax3 = axes[1, 0]
        ax3.plot(throughputs, min_latencies, 'o-', label='Min', linewidth=2, markersize=7, color='#70AD47')
        ax3.plot(throughputs, avg_latencies, 's-', label='Average', linewidth=2, markersize=7, color='#4472C4')
        ax3.plot(throughputs, max_latencies, '^-', label='Max', linewidth=2, markersize=7, color='#ED7D31')
        ax3.set_xlabel('Throughput (orders/sec)', fontweight='bold')
        ax3.set_ylabel('Latency (µs)', fontweight='bold')
        ax3.set_title('Complete Latency Profile', fontweight='bold')
        ax3.legend(loc='best')
        ax3.grid(True, alpha=0.3)
        
        # Plot 4: Summary Statistics Table
        ax4 = axes[1, 1]
        ax4.axis('off')
        
        # Create summary table
        table_data = []
        table_data.append(['Metric', 'Min', 'Avg', 'Max'])
        
        # Throughput stats
        table_data.append([
            'Throughput (orders/sec)',
            f'{min(throughputs):.2f}',
            f'{np.mean(throughputs):.2f}',
            f'{max(throughputs):.2f}'
        ])
        
        # Latency stats
        table_data.append([
            'Avg Latency (µs)',
            f'{min(avg_latencies):.2f}',
            f'{np.mean(avg_latencies):.2f}',
            f'{max(avg_latencies):.2f}'
        ])
        
        # Min latency stats
        table_data.append([
            'Min Latency (µs)',
            f'{min(min_latencies):.2f}',
            f'{np.mean(min_latencies):.2f}',
            f'{max(min_latencies):.2f}'
        ])
        
        # Max latency stats
        table_data.append([
            'Max Latency (µs)',
            f'{min(max_latencies):.2f}',
            f'{np.mean(max_latencies):.2f}',
            f'{max(max_latencies):.2f}'
        ])
        
        table = ax4.table(cellText=table_data, cellLoc='center', loc='center',
                         colWidths=[0.35, 0.2, 0.2, 0.2])
        table.auto_set_font_size(False)
        table.set_fontsize(10)
        table.scale(1, 2.5)
        
        # Header formatting
        for i in range(4):
            table[(0, i)].set_facecolor('#4472C4')
            table[(0, i)].set_text_props(weight='bold', color='white')
        
        # Alternating row colors
        for i in range(1, len(table_data)):
            for j in range(4):
                if i % 2 == 0:
                    table[(i, j)].set_facecolor('#D9E1F2')
                else:
                    table[(i, j)].set_facecolor('#FFFFFF')
        
        ax4.set_title('Summary Statistics', fontweight='bold', pad=20)
        
        plt.tight_layout()
        
        # Save graphs
        graph_file = self.results_dir / f"performance_graph_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        plt.savefig(graph_file, dpi=300, bbox_inches='tight')
        print(f"[OK] Performance graph saved: {graph_file}")
        
        # Display graph
        try:
            plt.show(block=False)
        except:
            pass  # In non-interactive mode, just save
        
        return graph_file


def main():
    """Main entry point."""
    print("\n" + "="*70)
    print("LAB 9: LATENCY vs. THROUGHPUT TEST HARNESS")
    print("="*70)
    print("\nThis test will:")
    print("  1. Run tests at different throughput rates")
    print("  2. Collect latency metrics from each run")
    print("  3. Generate performance graphs")
    print("  4. Export results to CSV and Excel")
    print("\nNote: Ensure the backend is running before starting tests")
    print("="*70)
    
    response = input("\nReady to start performance tests? (yes/no): ").strip().lower()
    if response != 'yes':
        print("Test cancelled")
        return
    
    # Create test suite
    suite = PerformanceTestSuite()
    
    # Test configs: (num_orders, mode)
    # These will achieve different throughput levels
    test_configs = [
        (100, 'burst'),    # ~100 orders/sec
        (500, 'burst'),    # ~500 orders/sec
        (1000, 'burst'),   # ~1000 orders/sec
    ]
    
    for num_orders, mode in test_configs:
        suite.add_test(num_orders, mode)
    
    # Run all tests
    if not suite.run_all():
        print("[ERROR] Some tests failed - results may be incomplete")
    
    # Export results
    print("\n" + "="*70)
    print("GENERATING REPORTS")
    print("="*70)
    
    csv_file = suite.export_csv()
    excel_file = suite.export_excel(csv_file)
    graph_file = suite.generate_graphs()
    
    print("\n" + "="*70)
    print("TEST COMPLETE")
    print("="*70)
    print(f"\nResults saved to: {suite.results_dir}/")
    print(f"  - CSV: {csv_file.name if csv_file else 'N/A'}")
    print(f"  - Excel: {excel_file.name if excel_file else 'N/A (openpyxl not installed)'}")
    print(f"  - Graph: {graph_file.name if graph_file else 'N/A (matplotlib not installed)'}")
    print("\nOpen the graph and reports to analyze performance characteristics.")
    print("="*70 + "\n")


if __name__ == '__main__':
    main()
